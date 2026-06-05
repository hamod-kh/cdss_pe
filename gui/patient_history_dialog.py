from __future__ import annotations
import os
from datetime import datetime

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QFrame,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QPushButton, QFileDialog, QMessageBox, QWidget
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor

import config.config as C
from gui.widgets import make_secondary_button, section_title_label, h_muted
from data.db import DatabaseHandler
from logger.logger import log_db_event, log_error

_COLUMNS = [
    ("Date / Time", 180),
    ("Wells Score", 100),
    ("Risk Level", 120),
    ("Key Data Summary", 240),
    ("Recommendation", 200),
    ("Decision", 120),
    ("Physician Comments", 220),
]


class PatientHistoryDialog(QDialog):

    def __init__(self, mrn: str, patient_name: str,
                 db: DatabaseHandler, parent=None) -> None:
        super().__init__(parent)
        self._mrn = mrn
        self._patient_name = patient_name
        self._db = db
        self._rows: list[list[str]] = []

        self.setWindowTitle(f"Patient History – {patient_name}  (MRN {mrn})")
        self.setMinimumSize(1100, 600)
        self.setModal(True)
        self.setStyleSheet("QDialog { background: white; } QLabel { background: transparent; }")

        self._build_ui()
        self._load_data()

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        root.addWidget(self._build_header())
        root.addWidget(self._build_table_area(), stretch=1)
        root.addWidget(self._build_footer())

    def _build_header(self) -> QFrame:
        bar = QFrame()
        bar.setStyleSheet(f"background: {C.COLOR_PRIMARY}; border-radius: 0px;")
        bar.setFixedHeight(52)
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(20, 0, 20, 0)
        title = QLabel("Assessment History")
        title.setStyleSheet("color: white; font-size: 15px; font-weight: 700; background: transparent;")
        layout.addWidget(title)
        layout.addStretch()
        sub = QLabel(f"{self._patient_name}  |  MRN {self._mrn}")
        sub.setStyleSheet("color: rgba(255,255,255,0.85); font-size: 12px; background: transparent;")
        layout.addWidget(sub)
        return bar

    def _build_table_area(self) -> QWidget:
        wrapper = QWidget()
        layout = QVBoxLayout(wrapper)
        layout.setContentsMargins(20, 16, 20, 8)
        layout.setSpacing(10)

        layout.addWidget(section_title_label("ASSESSMENT HISTORY"))

        div = QFrame()
        div.setFrameShape(QFrame.Shape.HLine)
        div.setFixedHeight(1)
        div.setStyleSheet(f"background: {C.COLOR_PRIMARY}; border: none; margin-bottom: 4px;")
        layout.addWidget(div)

        self._table = QTableWidget()
        self._table.setColumnCount(len(_COLUMNS))
        self._table.setHorizontalHeaderLabels([c[0] for c in _COLUMNS])
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setAlternatingRowColors(False)
        self._table.verticalHeader().setVisible(False)
        self._table.setShowGrid(True)
        self._table.setGridStyle(Qt.PenStyle.SolidLine)
        self._table.setStyleSheet(f"""
            QTableWidget {{
                border: 1px solid {C.COLOR_BORDER};
                border-radius: 6px;
                gridline-color: {C.COLOR_BORDER};
                font-size: 13px;
            }}
            QTableWidget::item {{
                padding: 8px 10px;
                border: none;
            }}
            QTableWidget::item:selected {{
                background-color: {C.COLOR_LIGHT_TEAL};
                color: #212529;
            }}
            QHeaderView::section {{
                background-color: {C.COLOR_PANEL_BG};
                padding: 8px 10px;
                font-weight: 700;
                font-size: 12px;
                border: none;
                border-bottom: 2px solid {C.COLOR_BORDER};
                border-right: 1px solid {C.COLOR_BORDER};
            }}
            QHeaderView::section:last {{
                border-right: none;
            }}
        """)

        header = self._table.horizontalHeader()
        for i, (_, width) in enumerate(_COLUMNS):
            if i in (3, 6):
                header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            else:
                header.setSectionResizeMode(i, QHeaderView.ResizeMode.Fixed)
                self._table.setColumnWidth(i, width)

        self._table.verticalHeader().setDefaultSectionSize(56)
        layout.addWidget(self._table)

        self._empty_lbl = QLabel("No assessments found for this patient.")
        self._empty_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty_lbl.setStyleSheet(f"color: {C.COLOR_TEXT_MUTED}; font-size: 13px; padding: 20px;")
        self._empty_lbl.setVisible(False)
        layout.addWidget(self._empty_lbl)

        return wrapper

    def _build_footer(self) -> QWidget:
        wrapper = QWidget()
        w_layout = QVBoxLayout(wrapper)
        w_layout.setContentsMargins(0, 0, 0, 0)
        w_layout.setSpacing(0)

        divider = QFrame()
        divider.setFrameShape(QFrame.Shape.HLine)
        divider.setFixedHeight(1)
        divider.setStyleSheet(f"background: {C.COLOR_BORDER}; border: none;")
        w_layout.addWidget(divider)

        bar = QFrame()
        bar.setStyleSheet(f"background: {C.COLOR_PANEL_BG};")
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(20, 10, 20, 10)
        layout.addWidget(h_muted("All times local"))
        layout.addStretch()
        layout.addWidget(h_muted(C.SOURCE_LABEL))
        layout.addSpacing(20)

        export_btn = make_secondary_button("Export to Excel")
        export_btn.clicked.connect(self._on_export)
        layout.addWidget(export_btn)

        close_btn = make_secondary_button("Close")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)

        w_layout.addWidget(bar)
        return wrapper

    def _load_data(self) -> None:
        records = self._db.assessment_history_full(self._mrn)
        self._rows.clear()
        self._table.setRowCount(0)

        if not records:
            self._table.setVisible(False)
            self._empty_lbl.setVisible(True)
            return

        self._table.setVisible(True)
        self._empty_lbl.setVisible(False)

        for rec in records:
            r = self._table.rowCount()
            self._table.insertRow(r)
            risk = rec.get("wells_risk", "")
            row_data = [
                rec.get("assessment_timestamp", ""),
                f"{rec.get('wells_score', 0.0):.1f}",
                risk,
                self._build_key_summary(rec),
                rec.get("recommendation_title", ""),
                rec.get("physician_decision", ""),
                rec.get("physician_comments", ""),
            ]
            self._rows.append(row_data)

            for col, val in enumerate(row_data):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                if col == 2:
                    if "HIGH" in val:
                        item.setForeground(QColor(C.COLOR_HIGH_RISK))
                        item.setFont(QFont("", -1, QFont.Weight.Bold))
                    elif "MODERATE" in val:
                        item.setForeground(QColor(C.COLOR_MODERATE))
                        item.setFont(QFont("", -1, QFont.Weight.Bold))
                    elif "LOW" in val:
                        item.setForeground(QColor(C.COLOR_LOW_RISK))
                        item.setFont(QFont("", -1, QFont.Weight.Bold))
                self._table.setItem(r, col, item)

        self._table.resizeRowsToContents()

    @staticmethod
    def _build_key_summary(rec: dict) -> str:
        parts: list[str] = []
        if rec.get("hr"):
            parts.append(f"HR {rec['hr']}")
        if rec.get("rr"):
            parts.append(f"RR {rec['rr']}")
        if rec.get("spo2"):
            parts.append(f"SpO2 {rec['spo2']}%")
        if rec.get("ddimer"):
            parts.append(f"D-dimer {rec['ddimer']} ng/mL")
        if rec.get("pleuritic_pain") == "Yes":
            parts.append("Pleuritic pain: Yes")
        if rec.get("dyspnea") == "Yes":
            parts.append("Dyspnea: Yes")
        return "  |  ".join(parts) if parts else "—"

    def _on_export(self) -> None:
        if not self._rows:
            QMessageBox.information(self, "No Data", "There is no data to export.")
            return
        default_name = f"PE_History_{self._mrn}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Export to Excel", default_name, "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            self._write_xlsx(path)
            log_db_event("EXPORT_HISTORY", f"MRN={self._mrn} -> {path}")
            QMessageBox.information(self, "Export Successful",
                                    f"History exported successfully to:\n{path}")
        except Exception as exc:
            log_error("PatientHistoryDialog._on_export", exc)
            QMessageBox.critical(self, "Export Failed", f"Could not write the file:\n{exc}")

    def _write_xlsx(self, path: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Assessment History"

        ws.merge_cells("A1:G1")
        ws["A1"] = f"Assessment History – {self._patient_name}  (MRN {self._mrn})"
        ws["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", start_color="0D7377")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   Source: CDS PE v1.0"
        ws["A2"].font = Font(name="Arial", size=10, italic=True, color="6C757D")
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18

        header_fill = PatternFill("solid", start_color="F8F9FA")
        header_font = Font(name="Arial", size=11, bold=True, color="212529")
        header_border = Border(
            bottom=Side(style="medium", color="0D7377"),
            right=Side(style="thin", color="DEE2E6"),
        )
        for col_i, (heading, _) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=3, column=col_i, value=heading)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 22

        thin_border = Border(
            bottom=Side(style="thin", color="DEE2E6"),
            right=Side(style="thin", color="DEE2E6"),
        )
        risk_colours = {
            "HIGH RISK": "C0392B",
            "MODERATE RISK": "E67E22",
            "LOW RISK": "27AE60",
        }
        for row_i, row_data in enumerate(self._rows, start=4):
            ws.row_dimensions[row_i].height = 36
            for col_i, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_i, column=col_i, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_i % 2 == 0:
                    cell.fill = PatternFill("solid", start_color="F8F9FA")
                if col_i == 3:
                    colour = risk_colours.get(value, "212529")
                    cell.font = Font(name="Arial", size=11, bold=True, color=colour)
                else:
                    cell.font = Font(name="Arial", size=11)

        col_widths = [22, 14, 16, 36, 28, 14, 32]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A4"
        wb.save(path)
